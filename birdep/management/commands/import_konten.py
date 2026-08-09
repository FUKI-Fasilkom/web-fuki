"""Impor konten BirDep, Program, Fungsionaris, dan Pengurus Inti dari file Excel.

Contoh pemakaian:

    # cek dulu tanpa menulis apa pun ke database
    python manage.py import_konten data/template-konten-fuki-2026.xlsx --dry-run

    # jalankan sungguhan, sekaligus memasang foto
    python manage.py import_konten data/template-konten-fuki-2026.xlsx --foto data/foto

Perintah ini aman dijalankan berulang kali. Baris yang sudah ada akan
diperbarui, bukan diduplikasi, sehingga foto susulan bisa dimasukkan kapan saja
tanpa mengacaukan data yang sudah masuk.
"""

from pathlib import Path

from django.core.files import File
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from birdep.models import BirDep, Fungsionaris, PengurusInti, Program

SHEET_WAJIB = ["birdep", "program", "fungsionaris", "pengurus_inti"]


def baca_sheet(ws, kolom_wajib):
    """Ubah satu sheet menjadi list of dict, memakai baris header pertama."""
    header_row = None
    for r in range(1, min(ws.max_row, 5) + 1):
        nilai = [str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        if all(k in nilai for k in kolom_wajib):
            header_row = r
            headers = nilai
            break
    if header_row is None:
        raise CommandError(
            f"Sheet '{ws.title}': tidak menemukan baris header yang memuat "
            f"kolom {kolom_wajib}"
        )

    hasil = []
    for r in range(header_row + 1, ws.max_row + 1):
        baris = {}
        for c, nama_kolom in enumerate(headers, start=1):
            if not nama_kolom:
                continue
            v = ws.cell(r, c).value
            baris[nama_kolom] = str(v).strip() if v is not None else ""
        if any(baris.get(k) for k in kolom_wajib):
            baris["_baris"] = r
            hasil.append(baris)
    return hasil


def as_int(nilai, default=0):
    try:
        return int(float(nilai))
    except (TypeError, ValueError):
        return default


class Command(BaseCommand):
    help = "Impor konten BirDep dan fungsionaris dari file Excel"

    def add_arguments(self, parser):
        parser.add_argument("berkas", help="Path ke file .xlsx")
        parser.add_argument(
            "--foto",
            dest="folder_foto",
            help="Folder berisi foto, dinamai sesuai kolom nama_file_foto",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Hanya memeriksa, tidak menulis apa pun ke database",
        )

    # ------------------------------------------------------------------ utils

    def info(self, teks):
        self.stdout.write(teks)

    def ok(self, teks):
        self.stdout.write(self.style.SUCCESS(teks))

    def warn(self, teks):
        self.stdout.write(self.style.WARNING(teks))

    def _indeks_foto(self):
        """Indeks isi folder foto, dikunci nama tanpa ekstensi dan huruf kecil.

        Dibuat sekali lalu dipakai ulang. Pencocokan sengaja mengabaikan beda
        huruf besar-kecil dan beda ekstensi, karena berkas seperti `123.JPG`
        lolos di Windows tapi tidak di server Linux.
        """
        if self._peta_foto is not None:
            return self._peta_foto
        self._peta_foto = {}
        if self.folder_foto:
            for berkas in sorted(self.folder_foto.iterdir()):
                if berkas.is_file():
                    self._peta_foto.setdefault(berkas.stem.lower(), berkas)
        return self._peta_foto

    def cari_foto(self, nama_berkas):
        """Kembalikan Path foto bila ada, jika tidak None."""
        if not (self.folder_foto and nama_berkas):
            return None
        return self._indeks_foto().get(Path(nama_berkas).stem.lower())

    def pasang_foto(self, obyek, nama_berkas):
        """Pasang foto ke obyek bila berkasnya ada dan belum terpasang."""
        if not nama_berkas:
            return "kosong"
        sumber = self.cari_foto(nama_berkas)
        if sumber is None:
            self.peringatan.append(f"foto tidak ditemukan: {nama_berkas} ({obyek.nama})")
            return "hilang"
        if obyek.foto and Path(obyek.foto.name).name == sumber.name:
            return "sudah ada"
        if not self.dry_run:
            with sumber.open("rb") as f:
                obyek.foto.save(sumber.name, File(f), save=True)
        return "dipasang"

    # ------------------------------------------------------------------ main

    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError("Paket openpyxl belum terpasang. Jalankan: pip install openpyxl") from exc

        berkas = Path(opts["berkas"])
        if not berkas.exists():
            raise CommandError(f"Berkas tidak ditemukan: {berkas}")

        self.dry_run = opts["dry_run"]
        self.folder_foto = Path(opts["folder_foto"]) if opts.get("folder_foto") else None
        if self.folder_foto and not self.folder_foto.is_dir():
            raise CommandError(f"Folder foto tidak ditemukan: {self.folder_foto}")

        self.peringatan = []
        self._peta_foto = None

        wb = openpyxl.load_workbook(berkas, data_only=True)
        kurang = [s for s in SHEET_WAJIB if s not in wb.sheetnames]
        if kurang:
            raise CommandError(f"Sheet berikut tidak ada di berkas: {', '.join(kurang)}")

        if self.dry_run:
            self.warn("MODE UJI COBA — tidak ada yang ditulis ke database\n")

        try:
            with transaction.atomic():
                self.impor_birdep(baca_sheet(wb["birdep"], ["slug", "nama"]))
                self.impor_program(baca_sheet(wb["program"], ["slug_birdep", "judul"]))
                self.impor_fungsionaris(
                    baca_sheet(wb["fungsionaris"], ["nama", "jabatan", "slug_birdep"])
                )
                self.impor_pengurus_inti(
                    baca_sheet(wb["pengurus_inti"], ["kategori", "nama", "jabatan"])
                )
                if self.dry_run:
                    raise _BatalkanUjiCoba()
        except _BatalkanUjiCoba:
            pass

        if self.peringatan:
            self.stdout.write("")
            self.warn(f"{len(self.peringatan)} peringatan:")
            for p in self.peringatan:
                self.warn(f"  - {p}")

        self.stdout.write("")
        if self.dry_run:
            self.ok("Pemeriksaan selesai. Jalankan ulang tanpa --dry-run untuk menyimpan.")
        else:
            self.ok("Impor selesai.")

    # --------------------------------------------------------------- bagian

    def impor_birdep(self, baris_list):
        dibuat = diperbarui = 0
        for b in baris_list:
            slug = b["slug"]
            if not slug:
                self.peringatan.append(f"birdep baris {b['_baris']}: slug kosong, dilewati")
                continue
            obyek, is_baru = BirDep.objects.get_or_create(
                slug=slug, defaults={"nama": b["nama"], "logo_filename": ""}
            )
            obyek.nama = b["nama"] or obyek.nama
            obyek.nama_panjang = b.get("nama_panjang", "")
            obyek.logo_filename = b.get("logo_filename", "")
            obyek.tentang_deskripsi = b.get("tentang_deskripsi", "")
            obyek.visi = b.get("visi", "")
            obyek.misi = b.get("misi", "")
            if not obyek.logo_filename:
                self.peringatan.append(f"birdep '{slug}': logo_filename kosong")
            if not obyek.tentang_deskripsi:
                self.peringatan.append(f"birdep '{slug}': tentang_deskripsi kosong")
            if not self.dry_run:
                obyek.save()
            dibuat += is_baru
            diperbarui += not is_baru
        self.info(f"BirDep         : {dibuat} baru, {diperbarui} diperbarui")

    def impor_program(self, baris_list):
        dibuat = diperbarui = dilewati = 0
        for b in baris_list:
            slug = b["slug_birdep"]
            birdep = BirDep.objects.filter(slug=slug).first()
            if birdep is None:
                self.peringatan.append(
                    f"program baris {b['_baris']}: slug_birdep '{slug}' tidak dikenal, dilewati"
                )
                dilewati += 1
                continue
            obyek, is_baru = Program.objects.get_or_create(
                birdep=birdep, judul=b["judul"], defaults={"deskripsi": ""}
            )
            obyek.deskripsi = b.get("deskripsi", "")
            obyek.urutan = as_int(b.get("urutan"))
            if not obyek.deskripsi:
                self.peringatan.append(f"program '{slug} / {b['judul']}': deskripsi kosong")
            if not self.dry_run:
                obyek.save()
            dibuat += is_baru
            diperbarui += not is_baru
        self.info(f"Program        : {dibuat} baru, {diperbarui} diperbarui, {dilewati} dilewati")

    def impor_fungsionaris(self, baris_list):
        dibuat = diperbarui = dilewati = 0
        foto_stat = {"dipasang": 0, "sudah ada": 0, "kosong": 0, "hilang": 0}
        for b in baris_list:
            slug = b["slug_birdep"]
            birdep = BirDep.objects.filter(slug=slug).first()
            if birdep is None:
                self.peringatan.append(
                    f"fungsionaris baris {b['_baris']}: slug_birdep '{slug}' tidak dikenal, dilewati"
                )
                dilewati += 1
                continue
            obyek, is_baru = Fungsionaris.objects.get_or_create(
                birdep=birdep, nama=b["nama"], defaults={"jabatan": b["jabatan"]}
            )
            obyek.jabatan = b["jabatan"]
            obyek.urutan = as_int(b.get("urutan"))
            if not self.dry_run:
                obyek.save()
            foto_stat[self.pasang_foto(obyek, b.get("nama_file_foto", ""))] += 1
            dibuat += is_baru
            diperbarui += not is_baru
        self.info(f"Fungsionaris   : {dibuat} baru, {diperbarui} diperbarui, {dilewati} dilewati")
        self.info(
            f"  foto         : {foto_stat['dipasang']} dipasang, "
            f"{foto_stat['sudah ada']} sudah ada, {foto_stat['kosong']} kosong, "
            f"{foto_stat['hilang']} tidak ketemu"
        )

    def impor_pengurus_inti(self, baris_list):
        dibuat = diperbarui = dilewati = 0
        foto_stat = {"dipasang": 0, "sudah ada": 0, "kosong": 0, "hilang": 0}
        sah = {k for k, _ in PengurusInti.KATEGORI_CHOICES}
        for b in baris_list:
            kategori = (b["kategori"] or "").lower()
            if kategori not in sah:
                self.peringatan.append(
                    f"pengurus_inti baris {b['_baris']}: kategori '{kategori}' "
                    f"tidak dikenal (pilihan: {', '.join(sorted(sah))}), dilewati"
                )
                dilewati += 1
                continue
            obyek, is_baru = PengurusInti.objects.get_or_create(
                kategori=kategori, nama=b["nama"], defaults={"jabatan": b["jabatan"]}
            )
            obyek.jabatan = b["jabatan"]
            obyek.urutan = as_int(b.get("urutan"))
            obyek.ikhtisar = b.get("ikhtisar", "")
            obyek.deskripsi_kerja = b.get("deskripsi_kerja", "")
            if not self.dry_run:
                obyek.save()
            foto_stat[self.pasang_foto(obyek, b.get("nama_file_foto", ""))] += 1
            dibuat += is_baru
            diperbarui += not is_baru
        self.info(f"Pengurus Inti  : {dibuat} baru, {diperbarui} diperbarui, {dilewati} dilewati")
        self.info(
            f"  foto         : {foto_stat['dipasang']} dipasang, "
            f"{foto_stat['sudah ada']} sudah ada, {foto_stat['kosong']} kosong, "
            f"{foto_stat['hilang']} tidak ketemu"
        )


class _BatalkanUjiCoba(Exception):
    """Dipakai untuk membatalkan transaksi saat mode uji coba."""
